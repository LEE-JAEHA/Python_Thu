import xlsxwriter
from openpyxl import load_workbook

load_wb = load_workbook("./test.xlsx",data_only=True)
# load_wb= load_wb.active
load_ws = load_wb['test']
load_ws2 = load_ws.columns
for row in load_ws:
    for cell in row:
        print(cell.value,end=" / ")
    print()
print("*"*10)
for cols in load_ws2:
    for j in cols:
        print(j.value,end=" / ")
    print()
print("*"*20)

load_ws.cell(row=7,column=7).value= 999

load_wb.save('test.xlsx')


