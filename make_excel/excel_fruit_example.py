import xlsxwriter
from openpyxl import load_workbook


load_wb = load_workbook("./fruit.xlsx",data_only=True)
load_ws = load_wb["Sheet1"]
load_ws.cell(row=1,column = 4).value ="총합"
for idx,row in enumerate(load_ws):
    if idx == 0:
        continue
    sum=1
    for idx2,data in enumerate(row):
        if idx2 ==0:
            continue
        elif idx2 ==1 or idx2 ==2:
            sum = sum * data.value
    load_ws.cell(row= idx+1,column=4).value = sum

load_wb.save("./fruit.xlsx")
load_wb.close()







# workbook = xlsxwriter.Workbook("hello.xlsx")
# worksheet = workbook.add_worksheet()
# for i in range(32):
#     worksheet.write(i,0,i+1)
# workbook.close()