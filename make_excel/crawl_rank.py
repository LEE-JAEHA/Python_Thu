from bs4 import BeautifulSoup
import requests
import xlsxwriter
from openpyxl import load_workbook

print("데이터 수집 시작")
target = "https://music.bugs.co.kr/chart"
raw = requests.get(target)

# print(raw.text)
html = BeautifulSoup(raw.text, "html.parser")
print(type(html))
title_list_tag = html.find_all("p", {"class":"title"} ) # get artist
artist_list_tag = html.find_all("p", {"class":"artist"} ) # get artist
# print(artist_list_tag)
title_list = list()
artist_list = list()
for i in title_list_tag:
    tmp = i.text
    tmp = tmp.replace("\n","",100)
    title_list.append(tmp)

for idx,v in enumerate(artist_list_tag):
    tmp = v.text
    if tmp.find(chr(13)) !=-1: # 아스키코드값 13 이라는 값이 들어 있다면... 이 사이트에서 가수가 2명이상 참여하면 뒤에 13 이라는 값을 넣어둠
        tmp = tmp.replace(chr(13), "", 10) # 그럼 그 부분을 아무것도 아닌걸로 놔둔다
        tmp=tmp[:len(tmp)//2] # 이것의 출력결과가 이승철이승철 이렇게 되기 때문에 이렇게 처리
    tmp = tmp.replace("\n", "", 100) # 엔터 개행 부분을 아무것도 아닌걸로 대체
    artist_list.append(tmp)
# #

#
for idx,val in enumerate(title_list):
    print("{0}위 / 제목 : {1} / 가수 : {2}".format(idx+1,val,artist_list[idx]))



load_wb = load_workbook("./rank.xlsx",data_only=True)
load_ws = load_wb['Sheet1']

load_ws.cell(row=1,column=1).value = "순위"
load_ws.cell(row=1,column=2).value = "제목"
load_ws.cell(row=1,column=3).value = "가수"


for idx,val in enumerate(title_list):
    load_ws.cell(row=idx+2, column=1).value = str(idx+1)+"위"
    load_ws.cell(row=idx+2, column=2).value = val
    load_ws.cell(row=idx+2, column=3).value = artist_list[idx]


load_wb.save('./rank.xlsx')
# artist = list()
# cnt = 1
# for i in artist_list_tag:
#     # print(i.get_text())
#     tmp = i.get_text()
#     tmp = tmp[1:]
#     rank = str(cnt) +". " + tmp
#     cnt += 1
#     artist.append(rank)
#
# for i in artist:
#     print(i)
